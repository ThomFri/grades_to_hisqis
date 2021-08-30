import os
import pandas as pandas
import xlwt
from xlrd import open_workbook
from openpyxl import load_workbook
import dateutil.parser
import numpy as np
from shutil import copyfile
from xlutils.copy import copy
import math
from enum import Enum

#gui
#https://stackoverflow.com/questions/45441885/how-can-i-create-a-dropdown-menu-from-a-list-in-tkinter
#https://www.delftstack.com/de/howto/python-tkinter/how-to-create-dropdown-menu-in-tkinter/
#https://docs.python.org/3/library/dialog.html
#https://pythonbasics.org/tkinter-filedialog/


# headers etc.
from python_modules.cache import read_cache, write_cache
from python_modules.config import read_config_sys_argv
from python_modules.input import get_input_int_config, file_selector_config, get_input_config
from python_modules.output import list_to_string_with_leading_index, print_program_title, save_file_selector


class Hdrs(Enum):
    MNR = "mtknr"
    ABS = "abschl"
    STG = "stg"
    NNA = "nachname"
    VNA = "vorname"
    BEW = "bewertung"
    PDA = "pdatum"
    RES = "res1"
    PST = "pstatus"
    POR = "pordnr"
    LNR = "labnr"

class Grd(Enum):
    NAN = "NAN"
    KAN = "KAN"

class Edg(Enum):
    START = "startHISsheet"
    END = "endHISsheet"

key = Hdrs.MNR
req_cols = {
        Hdrs.MNR: {
            "name": "mtknr",
            "nameNice": "Matrikelnummer"
        },
        Hdrs.BEW: {
            "name": "bewertung",
            "nameNice": "Bewertung"
        },
        Hdrs.PDA: {
            "name": "pdatum",
            "nameNice": "Prüfungsdatum",
            "special": "col/fixed"
        }
    }

use_easygui = True
use_file_picker = True

def find_in_workbook(wb, needle, skiprows=0):
    result = []
    sheet = wb.sheet_by_index(0)
    for rowidx in range(skiprows, sheet.nrows):
        row = sheet.row(rowidx)
        for colidx, cell in enumerate(row):
            if cell.value == needle:
                tmp_result = {
                    "col": colidx,
                    "row": rowidx
                }
                result.append(tmp_result)

    return result


def join_non_strings(join_string, iteratable):
    result = ""
    for item in iteratable:
        result += join_string + str(item)

    result = result[len(join_string):]

    return result


def clean_dataframe(df, cleaing_col, cleaning_set):
    to_delte = df[df[cleaing_col].isin(cleaning_set)].index
    df.drop(to_delte, inplace=True)
    df.reset_index(drop=True, inplace=True)

if __name__ == '__main__':
    print_program_title("GRADES ==TO==> HISQIS")
    print("")

    default_config_file = 'config.json'
    default_cache_file = "cache.json"
    config = read_config_sys_argv(default_config_file=default_config_file)
    cache = read_cache(cache_file=default_cache_file)


    hq_file = file_selector_config("Bitte von HisQis heruntergeladene Export-Datei (XLS) auswählen",
                                   config_item=config.get("hisqis_datei"),
                                   path=cache.get("last_hq_file"))
    cache["last_hq_file"] = hq_file
    print("\n")
    hq_wb = open_workbook(hq_file)
    tab_corners = {}
    tab_corners["start"] = find_in_workbook(hq_wb, Edg.START.value, skiprows=0)
    tab_corners["end"] = find_in_workbook(hq_wb, Edg.END.value, skiprows=0)


    skip_rows_till_tab = tab_corners["start"][0]["row"] + 1
    table_length = tab_corners["end"][1]["row"] - skip_rows_till_tab - 1
    hq_df = pandas.read_excel(hq_file, skiprows=skip_rows_till_tab, nrows=table_length)

    hq_key_name = key.value
    hq_default_key_col = 0
    if hq_key_name in hq_df.head():
        hq_df.set_index(hq_key_name)
        hq_index_col = hq_key_name
    else:
        hq_index_col = hq_df.columns[hq_default_key_col]
        hq_df.set_index([hq_index_col])

    hq_matrnr_set = set(hq_df[hq_index_col])

    own_file = file_selector_config("Bitte eigene Datei auswählen", show_files=False,
                                    config_item=config.get("eigene_datei"),
                                    path=os.path.dirname(hq_file))
    cache["last_own_file"] = own_file
    own_wb = load_workbook(own_file)
    own_wb_sheets = own_wb.sheetnames
    own_wb_sheet_name = 0

    if len(own_wb_sheets) > 1:
        print("\n")
        print("Ihre Excel-Datei enhält folgende Tabellenblätter: \n" + list_to_string_with_leading_index(own_wb_sheets))
        print("\n")
        own_wb_sheet_number = get_input_int_config("Welche Nummer (links) trägt das Tabellenblatt, das die Noten enhält?", range(len(own_wb_sheets)), config_item=config.get("arbeitsblatt_nummer"))
        own_wb_sheet_name = own_wb_sheets[own_wb_sheet_number]

    own_df = pandas.read_excel(own_file, header=None, sheet_name=own_wb_sheet_name)
    print("\n\n")


    print(own_df.head(10))

    print("\n")
    skip_rows_own = get_input_int_config("In welcher Zeilenzahl (links) befindet sich der Tabellenkopf in der oben angezeigten Tabelle?", config_item=config.get("eigene_zeilen_ueberspringen"))
    print("\n")

    own_df = pandas.read_excel(own_file, skiprows=skip_rows_own, sheet_name=own_wb_sheet_name)
    print("Die oben angezeigte Tabelle enhält folgende Spalten:")
    print(list_to_string_with_leading_index(own_df.columns))
    print("\n")

    print("Diese müssen Sie im Folgenden den Spalten von HisQis zuordnen.")
    print("\n")

    own_cols = {}
    fixed_values = {}

    col_config = config.get("eigene_spalten")
    for req_col in req_cols:
        special = req_cols[req_col].get("special")

        if not col_config is None:
            current_col_config = col_config.get(req_col.value)

            if current_col_config is None:
                current_col_config = {}

        else:
            current_col_config = {}

        ask_col = True

        if special == "col/fixed":
            special_input = get_input_int_config("Hat Ihre Tabelle eine " + req_cols[req_col]["nameNice"] + "-Spalte (in HisQis: \""+ req_cols[req_col]["name"] + "\")? [1 = ja, 0 = nein]", [0,1], config_item=current_col_config.get("spalte_vorhanden"))
            print("\n")

            if int(special_input) == 0:
                ask_col = False
                own_cols[req_col] = req_col

                fixed_values[req_col] = get_input_config("Bitte geben Sie den Festwert für " + req_cols[req_col]["nameNice"] + " (in HisQis: \""+ req_cols[req_col]["name"] + "\") ein:", config_item=current_col_config.get("festwert"))
                print("\n")

        if ask_col:
            col_index = get_input_int_config("Was ist die Nummer (links) Ihrer " + req_cols[req_col]["nameNice"] + "-Spalte (in HisQis: \""+ req_cols[req_col]["name"] + "\")?", range(len(own_df.columns)), config_item=current_col_config.get("spaltennummer"))
            print("\n")
            own_cols[req_col] = own_df.columns[col_index]

    print(own_df[own_df[own_cols[Hdrs.MNR]].notnull()].tail(10))

    last_rows_own = get_input_int_config("Bei welcher Zeilenzahl (links) endet der Inhalt der oben angezeigte Tabelle?", range(len(own_df)), config_item=config.get("eigene_ende"))
    
    nrows_own = last_rows_own + 1

    own_df = pandas.read_excel(own_file, skiprows=skip_rows_own, nrows=nrows_own, sheet_name=own_wb_sheet_name)

    #renaming own_df
    inverted_own_cols = {}
    for k, v in own_cols.items():
        inverted_own_cols[v] = k.value

    own_df = own_df.rename(columns=inverted_own_cols)
    own_df.set_index(key.value)
    if Hdrs.PDA in fixed_values:
        own_df[Hdrs.PDA.value] = fixed_values[Hdrs.PDA]


    own_matrnr_set = set(own_df[key.value])
    set_diff = hq_matrnr_set ^ own_matrnr_set

    if not len(set_diff) == 0:
        add_hq = hq_matrnr_set - own_matrnr_set
        add_own = own_matrnr_set - hq_matrnr_set

        print("\n\n")
        print("╔══════════════════════════════════════════════════════════════════════╗")
        print("║ WARNUNG!!!                                                           ║")
        print("║ ----------                                                           ║")
        print("║ Matrikelnummern stimmen nicht überein!                               ║")
        print("║                                                                      ║")
        print("║ Zusätzliche in HisQis-Datei: " + join_non_strings(", ",add_hq))
        print("║ Zusätzliche in eigener Datei: " + join_non_strings(", ",add_own))
        print("╚══════════════════════════════════════════════════════════════════════╝")
        print("\n")

        ignore_options = [
            "Nur die Matrikelnummern der HisQis-Datei berücksichtigen",
            "Nur die Matrikelnummern der eigenen Datei berücksichtigen",
            "Nur die Matrikelnummern berücksichtigen, die in beiden Dateien enthalten sind",
            "Die Matrikelnummern aus beiden Dateien berücksichtigen",
            "Mehr Details anzeigen"
        ]

        while True:
            do_loop = False
            do_ignore = get_input_int_config("Wie soll hiermit verfahren werden?" + "\n" +
                                      list_to_string_with_leading_index(ignore_options),
                                      range(len(ignore_options)),
                                      config_item=config.get("was_tun_wenn_matrikelnr_nicht_ueberein_stimmen")
                                      )

            if do_ignore == 0 or do_ignore == 2:
                clean_dataframe(own_df, key.value, add_own)
            if do_ignore == 1 or do_ignore == 2:
                clean_dataframe(hq_df, key.value, add_hq)
            if do_ignore == 4:
                print("\n\n")

                print("Zusätzlich in HisQis-Datei:")
                print("===========================")
                if len(add_hq) > 0:
                    with pandas.option_context('display.max_rows', None):  # more options can be specified also
                        to_print_rows = hq_df[hq_df[key.value].isin(add_hq)]
                        print(to_print_rows[[Hdrs.MNR.value, Hdrs.NNA.value, Hdrs.VNA.value, Hdrs.BEW.value, Hdrs.PST.value]])
                else:
                    print("Keine zusätzlichen Daten")

                print("\n\n")

                print("Zusätzlich in eigener Datei:")
                print("============================")
                if len(add_own) > 0:
                    with pandas.option_context('display.max_rows', None):
                        print(own_df[own_df[key.value].isin(add_own)])
                else:
                    print("Keine zusätzlichen Daten")

                print("\n\n")

                do_loop = True

            if not do_loop:
                break

    print("\n" + "Daten abgleichen..." + "\n")
    original_header = hq_df.columns
    #hq_df.drop(columns=[Hdrs.BEW.value, Hdrs.PDA.value], inplace=True)
    own_df = own_df[[key.value, Hdrs.BEW.value, Hdrs.PDA.value]]

    grade_mean = pandas.to_numeric(own_df[Hdrs.BEW.value], errors='coerce').mean()
    if grade_mean < 6:
        # needs multiplication!
        own_df[Hdrs.BEW.value] = own_df[Hdrs.BEW.value].apply(
            lambda x: x * 100
            if not isinstance(x, str)
            else x
        )

    merged_dataframe = pandas.merge(hq_df, own_df, on=key.value, how="outer")

    # just a helper
    def merger_non_nan(val1, val2):
        if pandas.isna(val1):
            return val2
        else:
            return val1


    merging_cols = [Hdrs.BEW.value, Hdrs.PDA.value]

    for merging_col in merging_cols:
        merged_dataframe[merging_col] = merged_dataframe.apply(
            lambda x: merger_non_nan(x[merging_col+"_x"], x[merging_col+"_y"]), axis=1
        )



    merged_dataframe = merged_dataframe[original_header]

    merged_dataframe[Hdrs.BEW.value] = merged_dataframe[Hdrs.BEW.value].apply(
        lambda x: x.upper()
        if isinstance(x,str)
        else x
    )



    bewertung_contains_nan = (merged_dataframe[Hdrs.BEW.value].isnull().sum() > 0)

    if bewertung_contains_nan:
        bewertung_options = [
            "Ignorieren",
            "Durch \""+Grd.NAN.value+"\" ersetzen",
            "Durch \""+Grd.KAN.value+"\" ersetzen"
        ]
        do_bewertung = get_input_int_config("Bewertung-Spalte (in HisQis: \"bewertung\") enthälte leere Werte! Wie soll mit diesen verfahren werden?" + "\n" +
                                     list_to_string_with_leading_index(bewertung_options),
                                     range(len(bewertung_options)),
                                     config_item=config.get("was_tun_wenn_bewertung_leere_werte_enthaelt")
                                     )
        print("\n")

        if do_bewertung == 1:
            merged_dataframe[Hdrs.BEW.value].replace(np.nan, Grd.NAN.value, regex=True, inplace=True)
        elif do_bewertung == 2:
            merged_dataframe[Hdrs.BEW.value].replace(np.nan, Grd.KAN.value, regex=True, inplace=True)

    merged_dataframe[Hdrs.PDA.value] = merged_dataframe[Hdrs.PDA.value].apply(
        lambda x: dateutil.parser.parse(str(x), dayfirst=True, yearfirst=False).strftime("%d.%m.%Y")
        if (np.all(pandas.notnull(x)))
        else x
    )

    do_target = get_input_int_config("Ergebnis direkt in HisQis-Datei schreiben? [1 = ja, 0 = nein / Kopie anlegen]", [0,1], config_item=config.get("in_hisqis_datei_schreiben"))
    if do_target == 0:
        last_dot = hq_file.rfind('.')
        target_file = hq_file[:last_dot] + "_upload" + hq_file[last_dot:]
        target_file_config = config.get("ziel_datei")

        if not target_file_config is None:
            target_file = target_file_config
        else:
            save_types = [["*.xls", "Excel Datei"]]

            last_target_file = cache.get("last_target_file")
            if last_target_file is None:
                last_target_file = target_file


            target_file = save_file_selector(text="Bitte Speicherziel für Upload-Datei auswählen",
                                             file_or_folder=last_target_file,
                                             gui=use_file_picker,
                                             save_filetypes=save_types)

        cache["last_target_file"] = os.path.basename(target_file)
        cache["last_folder_path"] = os.path.dirname(target_file)
        copyfile(hq_file, target_file)

    else:
        target_file = hq_file


    write_start = tab_corners["start"][0]["row"] + 1
    write_end_row = write_start + len(merged_dataframe) + 3
    write_end_col = tab_corners["end"][0]["col"] + 1

    target_wb_tmp = open_workbook(target_file, formatting_info=True)
    target_sheet_tmp = target_wb_tmp.sheet_by_index(0)
    target_wb = copy(target_wb_tmp)
    target_sheet = target_wb.get_sheet(0)
    nan_format = xlwt.XFStyle()
    nan_pattern = xlwt.Pattern()
    nan_pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    nan_pattern.pattern_fore_colour = xlwt.Style.colour_map['yellow']
    nan_format.pattern = nan_pattern

    row_i = write_start
    #for col, val in enumerate(merged_dataframe.columns, start=0):
    #    target_sheet.write(row_i, col, val)
    row_i += 1

    all_cols = list(merged_dataframe.columns)
    to_ignore_nans = [Hdrs.RES.value]
    to_ignore_col_indexes = []

    for to_ignore_nan in to_ignore_nans:
        tmp = all_cols.index(to_ignore_nan)
        to_ignore_col_indexes.append(tmp)


    for index, row_content in merged_dataframe.iterrows():
        for col, val in enumerate(row_content, start=0):
            if isinstance(val, float) and math.isnan(val):
                if col not in to_ignore_col_indexes:
                    target_sheet.write(row_i, col, style=nan_format)
            else:
                target_sheet.write(row_i, col, val)
        row_i += 1

    real_end_row = row_i
    real_end_col = tab_corners["start"][0]["col"]

    if write_end_row > row_i:
        for row_i in range(row_i, write_end_row):
            for col_i in range(0, write_end_col):
                target_sheet.write(row_i, col_i)

    target_sheet.write(real_end_row, real_end_col, Edg.END.value)

    target_wb.save(target_file)

    print("\n" + "FERTIG!")
    print("Sie können die Datei \"" + target_file + "\" jetzt auf HisQis hochladen.")
    print("Bitte kontrollieren Sie zuvor leere Zellen, die nicht gefüllt werden konnten; diese sind in gelb hervorgehoben")
    print("\n")

    do_open_file = get_input_int_config("Datei zur Kontrolle öffnen? [1 = ja, 0 = nein]", [0, 1], config_item=config.get("ziel_datei_oeffnen"))

    write_cache(cache, default_cache_file)

    if do_open_file == 1:
        os.startfile(os.path.normpath(target_file))
    pass


